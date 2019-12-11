from openpyxl import load_workbook

poll1 = load_workbook('poll1.xlsx')['1']
poll2 = load_workbook('poll2.xlsx')['1']

def make_dump(apps, schema_editor):
    Answer = apps.get_model('web', 'Answer')
    BigAnswer = apps.get_model('web', 'BigAnswer')
    Responder = apps.get_model('web', 'Responder')
    FirstPoll = apps.get_model('web', 'FirstPoll')
    SecondPoll = apps.get_model('web', 'SecondPoll')

    offensive_words_cell_pair = [
        ('B','C'),('D','E'),('F','G'),('H','I'),('J','K'),('L','M'),
        ('N','O'),('P','Q'),('R','S'),('T','U'),('V','W'),('X','Y'),
    ]

    free_time_cell_pair = [('Z','AA'),('AB','AC'),('AD','AE')]
    movie_cell_pair = [('AR','AS'),('AT','AU'),('AV','AW')]
    what_you_like_cell_pair = [('AF','AG'),('AH','AI'),('AJ','AK')]
    important_in_life_cell_pair = [('AX','AY'),('AZ','BA'),('BB','BC'),('BD','BD'),('BF','BG')]
    music_cell_pair = [('AL','AM'),('AN','AO'),('AP','AQ')]

    kurwa_cell_pair = [('B','C'), ('D','E'), ('F','G'), ('H','I'), ('J','K'), ('L','M')]
    chuj_cell_pair = [('O','P'), ('Q','R'), ('S','T'), ('U','V'), ('W','X'), ('Y','Z')]
    debil_cell_pair = [('AB','AC'), ('AD','AE'), ('AF','AG'), ('AH','AI'), ('AJ','AK'), ('AL','AM')]
    idiota_cell_pair = [('AO','AP'), ('AQ','AR'), ('AS','AT'), ('AU','AV')]
    szmata_cell_pair = [('AX','AY'), ('AZ','BA'), ('BB','BC'), ('BD','BE'), ('BF','BG')]
    pizda_cell_pair = [('BI','BJ'), ('BM','BN'), ('BO','BP'), ('BQ','BR'), ('BS','BT')]
    skuwysyn_cell_pair = [('BV','BW'), ('BX','BY'), ('BZ','CA'), ('CB','CC'), ('CD','CE')]
    dziwka_cell_pair = [('CG','CH'), ('CI','CJ'), ('CK','CL'), ('CM','CN'), ('CO','CP'), ('CQ','CR')]

    first_poll = FirstPoll()
    second_poll = SecondPoll()

    def answer_helper(poll, cell_pair_list, responder):
        answer_list = []
        for (value,category) in cell_pair_list:
            if value:
                value = poll[f'{value}{row}'].value
                category = poll[f'{category}{row}'].value
                answer = Answer(value=value, category=category, responder=responder)
                answer.save()
                answer_list.append(answer)
        return answer_list

    def big_answer_helper(poll, cell_pair_list, responder, rate_cell):
        big_answer = BigAnswer()
        answer_list = []
        for (value,category) in cell_pair_list:
            if value:
                value = poll[f'{value}{row}'].value
                category = poll[f'{category}{row}'].value
                answer = Answer(value=value, category=category, responder=responder)
                answer.save()
                answer_list.append(answer)
        big_answer.answers.set(answer_list)
        big_answer.power = poll[f'{rate_cell}{row}'].value
        big_answer.save()
        return big_answer

    for row in range(2,52):
        data = {
            "poll": 'poll1',
            "sex" : poll1[f'BH{row}'].value,
            "age" : poll1[f'BI{row}'].value,
            "education" : poll1[f'BJ{row}'].value,
            "living_place" : poll1[f'BK{row}'].value,
            "origin_place" : poll1[f'BL{row}'].value,
            "martial_status" : poll1[f'BM{row}'].value,
            "profession" : poll1[f'BN{row}'].value,
        }
        responder = Responder(**data)
        responder.save()

        offensive_words = []
        free_time = answer_helper(poll1, free_time_cell_pair, responder)
        movie = answer_helper(poll1, movie_cell_pair, responder)
        what_you_like = answer_helper(poll1, what_you_like_cell_pair, responder)
        important_in_life = answer_helper(poll1, important_in_life_cell_pair, responder)
        music = answer_helper(poll1, music_cell_pair, responder)

        for (value,power) in offensive_words_cell_pair:
            if value:
                value = poll1[f'{value}{row}'].value
                power = poll1[f'{power}{row}'].value
                word = Answer(value=value, power=power, responder=responder)
                word.save()
                offensive_word.append(word)

        first_poll.offensive_words.set(offensive_words)
        first_poll.free_time.set(free_time)
        first_poll.movie.set(movie)
        first_poll.what_you_like.set(what_you_like)
        first_poll.important_in_life.set(important_in_life)
        first_poll.save()



    for row in range(2,66):
        responder = Responder(poll="poll2")
        responder.save()

        kurwa = big_answer_helper(poll2, kurwa_cell_pair, responder, 'N')
        chuj = big_answer_helper(poll2, chuj_cell_pair, responder, 'AA')
        debil = big_answer_helper(poll2, debil_cell_pair, responder, 'AN')
        idiota = big_answer_helper(poll2, idiota_cell_pair, responder, 'AW')
        szmata = big_answer_helper(poll2, szmata_cell_pair, responder, 'BH')
        pizda = big_answer_helper(poll2, pizda_cell_pair, responder, 'BU')
        skuwysyn = big_answer_helper(poll2, skuwysyn_cell_pair, responder, 'CF')
        dziwka = big_answer_helper(poll2, dziwka_cell_pair, responder, 'CS')

        second_poll.kurwa.set(kurwa)
        second_poll.chuj.set(chuj)
        second_poll.debil.set(movie)
        second_poll.idiota.set(idiota)
        second_poll.szmata.set(szmata)
        second_poll.pizda.set(pizda)
        second_poll.skurwysyn.set(skurwysyn)
        second_poll.dziwka.set(dziwka)
        second_poll.save()
